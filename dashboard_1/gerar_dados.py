import json
import math
import pandas as pd
import warnings
from prophet import Prophet
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# Ignora avisos de tipos mistos que podem aparecer ao ler arquivos CSV grandes
warnings.filterwarnings("ignore")


class ComparadorEscalacao:
    def __init__(self, siacesp_path, data_posicao_alvo, mes_etb_alvo=5):
        """
        Inicializa a classe centralizando a base do SIACESP.
        """
        self.df_siacesp = pd.read_excel(siacesp_path)
        self.data_posicao_alvo = pd.to_datetime(data_posicao_alvo).date()
        self.mes_etb_alvo = mes_etb_alvo

        # 1. Converte a coluna de data do SIACESP (ajuste 'Z_PERÍODO' se o nome da coluna for diferente)
        self.df_siacesp["Z_PERÍODO"] = pd.to_datetime(
            self.df_siacesp["Z_PERÍODO"], errors="coerce"
        )

        # 2. Filtra o SIACESP para o mês e ano desejados (Ex: Maio de 2026)
        df_siacesp_filtrado = self.df_siacesp[
            (self.df_siacesp["Z_PERÍODO"].dt.month == self.mes_etb_alvo)
            & (self.df_siacesp["Z_PERÍODO"].dt.year == 2026)
        ]

        # 3. Converte o volume para numérico e agrega APENAS a base filtrada
        df_siacesp_filtrado["VOLUME"] = pd.to_numeric(
            df_siacesp_filtrado["VOLUME"], errors="coerce"
        ).fillna(0)
        self.siacesp_agg = df_siacesp_filtrado.groupby(
            "Y_PRODUTO PARA", as_index=False
        )["VOLUME"].sum()
        self.siacesp_agg.rename(columns={"VOLUME": "Volume_SIACESP"}, inplace=True)

    def processar_lineup(self, nome, caminho, col_data_pos, col_etb, col_volume):
        """
        Lê, limpa, filtra e compara um arquivo individual de Escalação (Line Up).
        """
        df = pd.read_excel(caminho)

        # Tratamento das colunas de data
        df[col_data_pos] = pd.to_datetime(df[col_data_pos], errors="coerce")
        # format='mixed' e dayfirst=False ajudam com padrões irregulares como MM/DD/YYYY
        df[col_etb] = pd.to_datetime(
            df[col_etb], errors="coerce", format="mixed", dayfirst=False
        )

        # Aplica o filtro de Data da Posição e Mês do ETB
        df_filtrado = df[df[col_data_pos].dt.date == self.data_posicao_alvo]
        df_mes = df_filtrado[df_filtrado[col_etb].dt.month == self.mes_etb_alvo]

        # Converte a coluna de volume customizada e agrega os dados
        df_mes[col_volume] = pd.to_numeric(df_mes[col_volume], errors="coerce").fillna(
            0
        )
        lineup_agg = df_mes.groupby("Y_PRODUTO PARA", as_index=False)[col_volume].sum()
        lineup_agg.rename(columns={col_volume: "Volume_LineUp"}, inplace=True)

        # Realiza um outer merge para garantir que produtos exclusivos de uma base também apareçam
        comparacao = pd.merge(
            lineup_agg, self.siacesp_agg, on="Y_PRODUTO PARA", how="outer"
        ).fillna(0)

        # Calcula a discrepância volumétrica pura
        comparacao["Discrepancia"] = (
            comparacao["Volume_LineUp"] - comparacao["Volume_SIACESP"]
        )

        # Calcula o % DISC individual linha a linha
        def calc_perc(row):
            if row["Volume_SIACESP"] > 0:
                return row["Discrepancia"] / row["Volume_SIACESP"]
            elif row["Volume_LineUp"] > 0:
                return (
                    1.0  # 100% de erro se não havia no SIACESP mas apareceu no Line Up
                )
            return 0.0

        comparacao["% DISC."] = comparacao.apply(calc_perc, axis=1)

        # Organiza as maiores discrepâncias no topo e adiciona a identificação da origem
        comparacao = comparacao.sort_values(
            by="Discrepancia", key=abs, ascending=False
        ).reset_index(drop=True)
        comparacao.insert(0, "Fonte Line Up", nome)

        # =======================================================
        # LÓGICA WAPE: CÁLCULO DA LINHA DE TOTAL PONDERADA
        # =======================================================
        total_lineup = comparacao["Volume_LineUp"].sum()
        total_siacesp = comparacao["Volume_SIACESP"].sum()
        diferenca_liquida = total_lineup - total_siacesp

        # O pulo do gato: Somar os erros absolutos para que não se anulem
        soma_erros_absolutos = comparacao["Discrepancia"].abs().sum()
        wape_percentual = (
            (soma_erros_absolutos / total_siacesp) if total_siacesp > 0 else 0
        )

        linha_total = pd.DataFrame(
            [
                {
                    "Fonte Line Up": nome,
                    "Y_PRODUTO PARA": "TOTAL",
                    "Volume_LineUp": total_lineup,
                    "Volume_SIACESP": total_siacesp,
                    "Discrepancia": diferenca_liquida,
                    "% DISC.": wape_percentual,
                }
            ]
        )

        # Anexa a linha de total formatada ao fim do DataFrame
        comparacao = pd.concat([comparacao, linha_total], ignore_index=True)

        # Arredonda os valores decimais para manter o relatório limpo
        comparacao["Volume_LineUp"] = comparacao["Volume_LineUp"].round(2)
        comparacao["Volume_SIACESP"] = comparacao["Volume_SIACESP"].round(2)
        comparacao["Discrepancia"] = comparacao["Discrepancia"].round(2)
        comparacao["% DISC."] = comparacao["% DISC."].round(4)

        return comparacao


def aplicar_estilo_planilha(ws, titulo):
    """Aplica estilos a uma planilha do Excel"""
    header_fill = PatternFill(
        start_color="161B22", end_color="161B22", fill_type="solid"
    )
    header_font = Font(bold=True, color="E6EDF3")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")


def gerar_dados_previsao(
    df_siacesp,
    coluna_data="Z_PERÍODO",
    coluna_volume="VOLUME",
    coluna_produto="Y_PRODUTO PARA",
    meses_futuros=6,
):
    """
    ✅ CORRIGIDO: Treina múltiplos modelos Prophet (Geral e Específicos por Produto)
    e exporta um dicionário segmentado com VALIDAÇÃO COMPLETA.
    """
    df = df_siacesp.copy()
    df["ds"] = pd.to_datetime(df[coluna_data], errors="coerce")
    df = df.dropna(subset=["ds"])
    df[coluna_volume] = pd.to_numeric(df[coluna_volume], errors="coerce").fillna(0)

    # Lista de produtos alvo + cenário geral
    produtos_alvo = ["MOP", "SAM", "UREA", "MAP", "NP", "SSP", "TSP"]
    resultados_exportacao = {}

    print("Iniciando geração de previsões com Prophet...")

    # Função interna para treinar o Prophet e evitar repetição de código
    def processar_modelo(df_filtrado, nome_chave):
        print(f"  ⏳ Processando modelo: {nome_chave}...")

        # Se não houver dados para o produto, retorna lista vazia
        if df_filtrado.empty:
            print(f"    ⚠️  Sem dados para: {nome_chave}")
            resultados_exportacao[nome_chave] = []
            return

        df_agg = (
            df_filtrado.groupby(df_filtrado["ds"].dt.to_period("M"))[coluna_volume]
            .sum()
            .reset_index()
        )
        df_agg["ds"] = df_agg["ds"].dt.to_timestamp()
        df_agg.rename(columns={coluna_volume: "y"}, inplace=True)

        # Prophet precisa de um histórico mínimo (ex: 2 meses) para não quebrar
        if len(df_agg) < 2:
            print(f"    ⚠️  Histórico insuficiente (< 2 meses) para: {nome_chave}")
            resultados_exportacao[nome_chave] = []
            return

        try:
            modelo = Prophet(
                yearly_seasonality=True,
                weekly_seasonality=False,
                daily_seasonality=False,
            )
            modelo.fit(df_agg)

            futuro = modelo.make_future_dataframe(periods=meses_futuros, freq="MS")
            previsao = modelo.predict(futuro)

            dados_js = []
            for index, row in previsao.iterrows():
                data_str = row["ds"].strftime("%Y-%m")

                real_row = df_agg[df_agg["ds"] == row["ds"]]
                valor_real_numpy = (
                    real_row["y"].values[0] if not real_row.empty else None
                )

                valor_real = (
                    None
                    if pd.isna(valor_real_numpy)
                    else round(float(valor_real_numpy), 2)
                )

                dados_js.append(
                    {
                        "data": data_str,
                        "real": valor_real,
                        "tendencia": round(float(row["yhat"]), 2),
                        "limite_inferior": round(float(row["yhat_lower"]), 2),
                        "limite_superior": round(float(row["yhat_upper"]), 2),
                    }
                )

            resultados_exportacao[nome_chave] = dados_js
            print(f"    ✅ {nome_chave}: {len(dados_js)} pontos de dados gerados")

        except Exception as e:
            print(f"    ❌ Erro ao treinar {nome_chave}: {str(e)}")
            resultados_exportacao[nome_chave] = []

    # 1. Roda o modelo GERAL (com todos os dados)
    processar_modelo(df, "GERAL")

    # 2. Roda um modelo para CADA PRODUTO específico
    for produto in produtos_alvo:
        # Filtra ignorando maiúsculas/minúsculas e garantindo que é texto
        df_produto = df[df[coluna_produto].fillna("").str.upper() == produto.upper()]
        processar_modelo(df_produto, produto)

    print(f"\n✅ Previsões geradas com sucesso!")
    print(
        f"   Produtos com dados: {sum(1 for v in resultados_exportacao.values() if v)}"
    )

    return resultados_exportacao


def exportar_dados_para_html(
    excel_path, df_siacesp_bruto=None, dict_lineups=None, config_sop=None
):
    """
    Lê o arquivo Excel gerado, consolida os dados raw (SIACESP + LineUps)
    e exporta para um arquivo JavaScript para o Dashboard.
    """
    try:
        xls = pd.ExcelFile(excel_path)

        def ler_aba(nome_aba):
            if nome_aba in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=nome_aba, skiprows=2)
                if "Y_PRODUTO PARA" in df.columns:
                    df = df[df["Y_PRODUTO PARA"] != "TOTAL"]

                return df.fillna(0).to_dict(orient="records")
            return []

        if isinstance(df_siacesp_bruto, str):
            df_siacesp_bruto = pd.read_excel(df_siacesp_bruto)

        if df_siacesp_bruto is not None:
            for col in df_siacesp_bruto.select_dtypes(
                include=["datetime64", "datetimetz"]
            ).columns:
                df_siacesp_bruto[col] = df_siacesp_bruto[col].astype(str)

        # --- CONSOLIDADOR DE MÚLTIPLOS LINE UPS ---
        lineup_raw_consolidado = []
        if dict_lineups:
            for nome_fonte, caminho_ou_df in dict_lineups.items():
                if isinstance(caminho_ou_df, str):
                    df_temp = pd.read_excel(caminho_ou_df)
                else:
                    df_temp = caminho_ou_df.copy()

                for col in df_temp.select_dtypes(
                    include=["datetime64", "datetimetz"]
                ).columns:
                    df_temp[col] = df_temp[col].astype(str)

                df_temp["FONTE_DADOS"] = nome_fonte  # Carimba de onde veio a informação
                lineup_raw_consolidado.extend(
                    df_temp.fillna(0).to_dict(orient="records")
                )

        if not config_sop and dict_lineups:
            todas_fontes = list(dict_lineups.keys())
            config_sop = {1: todas_fontes, 2: todas_fontes, 3: todas_fontes}

        datasets = {
            "resumo": ler_aba("Resumo Consolidado"),
            "unimar": ler_aba("Comp_UNIMAR"),
            "transatlantica": ler_aba("Comp_TRANSATLANTICA"),
            "wilson": ler_aba("Comp_WILSON_SONS"),
            "orion": ler_aba("Comp_ORION"),
            "siacesp_base": ler_aba("Base_SIACESP_Agregada"),
            "siacesp_raw": (
                df_siacesp_bruto.fillna(0).to_dict(orient="records")
                if df_siacesp_bruto is not None
                else []
            ),
            "lineup_raw": lineup_raw_consolidado,
            "config_sop": config_sop or {},
        }

        with open("dados_dashboard.js", "w", encoding="utf-8") as f:
            f.write("const dashboardData = " + json.dumps(datasets, default=str) + ";")

        print("✅ Arquivo 'dados_dashboard.js' atualizado com Suporte Multi-Lineup!")
    except Exception as e:
        print(f"❌ Erro ao exportar dados para HTML: {e}")


def main():
    print("=" * 60)
    print("PIPELINE COMPLETO: Escalação + Previsão com Prophet")
    print("=" * 60 + "\n")

    # 1. Instanciar a classe base configurando Data da Posição (2026-06-11) e Mês ETB (Maio)
    print("1️⃣  Carregando dados SIACESP...")
    comparador = ComparadorEscalacao("BI_Importacao Siacesp.xlsx", "2026-06-11", 5)
    print(f"   ✅ SIACESP carregado: {len(comparador.df_siacesp)} registros\n")

    # 2. Processar cada base com suas colunas de volume específicas
    print("2️⃣  Processando Line Ups...")
    comp_unimar = comparador.processar_lineup(
        nome="UNIMAR",
        caminho="BI_Line Up.xlsx",
        col_data_pos="DATA_POSIÇÃO",
        col_etb="ETB",
        col_volume="TONNAGE",
    )
    print(f"   ✅ UNIMAR processado: {len(comp_unimar)} linhas")

    comp_trans = comparador.processar_lineup(
        nome="TRANSATLANTICA",
        caminho="BI_Line Up_TRANSATLANTICA.xlsx",
        col_data_pos="DATA POSIÇÃO",
        col_etb="ETB",
        col_volume="QUANTITY",
    )
    print(f"   ✅ TRANSATLANTICA processado: {len(comp_trans)} linhas")

    comp_wilson = comparador.processar_lineup(
        nome="WILSON SONS",
        caminho="BI_Line Up_WILSON SONS.xlsx",
        col_data_pos="DATA POSIÇÃO",
        col_etb="ETB",
        col_volume="WEIGHT",
    )
    print(f"   ✅ WILSON SONS processado: {len(comp_wilson)} linhas")

    comp_orion = comparador.processar_lineup(
        nome="ORION",
        caminho="BI_Line Up_ORION.xlsx",
        col_data_pos="DATA POSIÇÃO",
        col_etb="Z_ETB",
        col_volume="Qtty",
    )
    print(f"   ✅ ORION processado: {len(comp_orion)} linhas\n")

    # 3. Concatenar as bases e extrair as Top 20 maiores divergências num contexto consolidado
    print("3️⃣  Consolidando dados...")
    resumo = pd.concat([comp_unimar, comp_trans, comp_wilson, comp_orion])

    # IMPORTANTE: A linha de "TOTAL" agora é excluída do 'resumo' para não entrar no ranking das Top 20
    resumo = (
        resumo[(resumo["Discrepancia"] != 0) & (resumo["Y_PRODUTO PARA"] != "TOTAL")]
        .sort_values(by="Discrepancia", key=abs, ascending=False)
        .head(20)
    )
    print(f"   ✅ Top 20 divergências identificadas\n")

    # 4. Construção do Arquivo Excel usando OpenPyXL
    print("4️⃣  Gerando arquivo Excel...")
    wb = Workbook()

    # Aba 1: Resumo Global
    ws_summary = wb.active
    ws_summary.title = "Resumo Consolidado"
    for r in dataframe_to_rows(resumo, index=False, header=True):
        ws_summary.append(r)
    aplicar_estilo_planilha(
        ws_summary, "Top 20 Discrepâncias - Múltiplas Fontes (Maio 2026)"
    )

    # Aba 2: UNIMAR
    ws_unimar = wb.create_sheet(title="Comp_UNIMAR")
    for r in dataframe_to_rows(comp_unimar, index=False, header=True):
        ws_unimar.append(r)
    aplicar_estilo_planilha(ws_unimar, "Comparação: Line Up UNIMAR vs SIACESP")

    # Aba 3: Transatlântica
    ws_trans = wb.create_sheet(title="Comp_TRANSATLANTICA")
    for r in dataframe_to_rows(comp_trans, index=False, header=True):
        ws_trans.append(r)
    aplicar_estilo_planilha(ws_trans, "Comparação: Line Up TRANSATLANTICA vs SIACESP")

    # Aba 4: Wilson Sons
    ws_wilson = wb.create_sheet(title="Comp_WILSON_SONS")
    for r in dataframe_to_rows(comp_wilson, index=False, header=True):
        ws_wilson.append(r)
    aplicar_estilo_planilha(ws_wilson, "Comparação: Line Up WILSON SONS vs SIACESP")

    # Aba 5: ORION
    ws_orion = wb.create_sheet(title="Comp_ORION")
    for r in dataframe_to_rows(comp_orion, index=False, header=True):
        ws_orion.append(r)
    aplicar_estilo_planilha(ws_orion, "Comparação: Line Up ORION vs SIACESP")

    # Aba 6: Base de Segurança SIACESP
    ws_siacesp = wb.create_sheet(title="Base_SIACESP_Agregada")
    for r in dataframe_to_rows(
        comparador.siacesp_agg.sort_values(by="Volume_SIACESP", ascending=False),
        index=False,
        header=True,
    ):
        ws_siacesp.append(r)
    aplicar_estilo_planilha(ws_siacesp, "SIACESP - Volume Total por Produto")

    # 5. Salva o resultado final
    wb.save("Analise_Comparativa_Escalacao.xlsx")
    print("   ✅ Excel gerado: 'Analise_Comparativa_Escalacao.xlsx'\n")

    # 6. EXPORTAÇÃO PARA HTML
    print("5️⃣  Exportando dados para Dashboard...")
    exportar_dados_para_html(
        excel_path="Analise_Comparativa_Escalacao.xlsx",
        df_siacesp_bruto="BI_Importacao Siacesp.xlsx",
        dict_lineups={
            "ORION": "BI_Line Up_ORION.xlsx",
            "TRANSATLANTICA": "BI_Line Up_TRANSATLANTICA.xlsx",
            "WILSON SONS": "BI_Line Up_WILSON SONS.xlsx",
            "UNIMAR": "BI_Line Up.xlsx",
        },
        config_sop={
            1: ["ORION"],
            2: ["ORION"],
            3: ["ORION"],
        },
    )
    print("   ✅ Dados consolidados exportados\n")

    # 7. GERAÇÃO DE PREVISÕES PROPHET ✅
    print("6️⃣  Gerando previsões com Prophet...")
    dados_prophet = gerar_dados_previsao(
        df_siacesp=comparador.df_siacesp,
        coluna_produto="Y_PRODUTO PARA",
        coluna_volume="VOLUME",
        meses_futuros=7,
    )

    # VALIDAÇÃO: Garante que pelo menos GERAL tem dados
    if "GERAL" not in dados_prophet or not dados_prophet["GERAL"]:
        print("❌ ERRO: Modelo GERAL não gerou dados! Verifique o histórico de dados.")
        return

    with open("dados_dashboard.js", "a", encoding="utf-8") as f:
        f.write(f"\n\nconst FORECAST_DATA = {json.dumps(dados_prophet, indent=2)};\n")

    print(
        f"   ✅ Dados Prophet exportados: {json.dumps({k: len(v) for k, v in dados_prophet.items()}, indent=6)}\n"
    )

    print("=" * 60)
    print("✅ PIPELINE COMPLETO COM SUCESSO!")
    print("=" * 60)
    print("\n Arquivos gerados:")
    print("   • Analise_Comparativa_Escalacao.xlsx (Relatório executivo)")
    print("   • dados_dashboard.js (Base de dados do Dashboard)")


# Executa o pipeline
if __name__ == "__main__":
    main()
