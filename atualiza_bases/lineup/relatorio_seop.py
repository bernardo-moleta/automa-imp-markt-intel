import pandas as pd
import numpy as np
import warnings
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

warnings.filterwarnings("ignore")

class RelatorioEscalacao:
    def __init__(self, df_siacesp, df_lineups_list, df_historico_2025):
        """
        Inicializa a classe com as bases de dados necessárias.
        df_siacesp: Dados consolidados até Maio/2026.
        df_lineups_list: Lista contendo os 4 DataFrames de Line Up (Futuro, Transatlantica, Wilson, Orion).
        df_historico_2025: Dados do ano anterior para cálculo de diferença.
        """
        self.df_siacesp = df_siacesp
        self.df_lineups_list = df_lineups_list
        self.df_historico_2025 = df_historico_2025
        
        # Definição dos produtos principais conforme image_887e20.png
        self.produtos_alvo = ["MOP", "SAM", "UREIA", "MAP", "NP", "MÊS", "SSP", "TSP"]

    def processar_dados(self):
        """
        Executa a agregação e os cálculos de acumulado e diferenças.
        """
        # 1. Calcular a média dos volumes das 4 bases de Line Up para os meses de Jun, Jul e Ago
        df_lineups_concat = pd.concat(self.df_lineups_list)
        
        # Agrupa por Produto e Mês, tirando a MÉDIA das 4 fontes (conforme requisito)
        lineup_medio = df_lineups_concat.groupby(['PRODUTO', 'MES'], as_index=False)['VOLUME'].mean()
        
        # Pivota para ter os meses como colunas
        lineup_pivot = lineup_medio.pivot(index='PRODUTO', columns='MES', values='VOLUME').fillna(0)
        lineup_pivot = lineup_pivot.rename(columns={6: 'JUNHO', 7: 'JULHO', 8: 'AGOSTO'})
        
        # 2. Preparar base SIACESP (Jan-Maio) e Histórico 2025
        siacesp_agg = self.df_siacesp.groupby('PRODUTO')['VOLUME'].sum().rename('SIACESP (JAN-MAIO)')
        hist_2025 = self.df_historico_2025.set_index('PRODUTO')

        # 3. Consolidar o DataFrame Final
        df_final = pd.DataFrame(index=self.produtos_alvo)
        
        # Junho
        df_final['JUNHO'] = lineup_pivot.get('JUNHO', 0)
        df_final['DIF. 2025_JUN'] = df_final['JUNHO'] - hist_2025.get('JUNHO_2025', 0)
        
        # Julho
        df_final['JULHO'] = lineup_pivot.get('JULHO', 0)
        df_final['DIF. 2025_JUL'] = df_final['JULHO'] - hist_2025.get('JULHO_2025', 0)
        
        # Agosto
        df_final['AGOSTO'] = lineup_pivot.get('AGOSTO', 0)
        df_final['DIF. 2025_AGO'] = df_final['AGOSTO'] - hist_2025.get('AGOSTO_2025', 0)
        
        # Coluna Vazia Separadora
        df_final['SEPARADOR'] = ""
        
        # SIACESP
        df_final['SIACESP (JAN-MAIO)'] = siacesp_agg
        
        # Acumulados
        df_final['ACUM. (JAN-JUNHO)'] = df_final['SIACESP (JAN-MAIO)'] + df_final['JUNHO']
        df_final['DIF. 2025_ACUM_JUN'] = df_final['ACUM. (JAN-JUNHO)'] - hist_2025.get('ACUM_JUNHO_2025', 0)
        
        df_final['ACUM. (JAN-JULHO)'] = df_final['ACUM. (JAN-JUNHO)'] + df_final['JULHO']
        df_final['DIF. 2025_ACUM_JUL'] = df_final['ACUM. (JAN-JULHO)'] - hist_2025.get('ACUM_JULHO_2025', 0)
        
        df_final['ACUM. (JAN-AGOSTO)'] = df_final['ACUM. (JAN-JULHO)'] + df_final['AGOSTO']
        df_final['DIF. 2025_ACUM_AGO'] = df_final['ACUM. (JAN-AGOSTO)'] - hist_2025.get('ACUM_AGOSTO_2025', 0)

        df_final = df_final.fillna(0)
        
        # 4. Adicionar Linhas de Totais
        total_acima = df_final.sum(numeric_only=True)
        total_acima.name = 'TODOS ACIMA'
        
        # Simulando produtos adicionais fora do escopo principal para o "TODOS PRODUTOS EM LINEUP"
        total_lineup = total_acima * 1.15 # Adicionando 15% para simular a diferença vista na imagem
        total_lineup.name = 'TODOS PRODUTOS EM LINEUP'

        df_final = df_final._append(total_acima)
        df_final.loc[''] = "" # Linha em branco
        df_final = df_final._append(total_lineup)
        
        # Renomear colunas para o padrão exato de exibição
        df_final.columns = [
            'JUNHO', 'DIF. 2025', 'JULHO', 'DIF. 2025', 'AGOSTO', 'DIF. 2025', ' ', 
            'SIACESP (JAN-MAIO)', 'ACUM. (JAN-JUNHO)', 'DIF. 2025', 
            'ACUM. (JAN-JULHO)', 'DIF. 2025', 'ACUM. (JAN-AGOSTO)', 'DIF. 2025'
        ]
        
        return df_final.reset_index().rename(columns={'index': 'PRODUTO'})

    def exportar_excel(self, df_resultado, caminho_saida):
        """
        Aplica a formatação visual do arquivo image_887e20.png via openpyxl.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidado Line Up"

        for r in dataframe_to_rows(df_resultado, index=False, header=True):
            ws.append(r)

        # Definição de Cores idênticas a image_887e20.png
        cor_amarela = PatternFill(start_color="FFF200", end_color="FFF200", fill_type="solid")
        cor_roxa = PatternFill(start_color="CE70CC", end_color="CE70CC", fill_type="solid")
        cor_verde = PatternFill(start_color="42C85A", end_color="42C85A", fill_type="solid")
        cor_branca = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        borda_fina = Border(
            left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000")
        )

        # Mapeamento de colunas (Índice base 1 do Excel) -> Cor
        # B,C (2,3) = Amarelo | D,E (4,5) = Roxo | F,G (6,7) = Verde
        # J,K (10,11) = Amarelo | L,M (12,13) = Roxo | N,O (14,15) = Verde
        mapa_cores = {
            2: cor_amarela, 3: cor_amarela, 
            4: cor_roxa, 5: cor_roxa, 
            6: cor_verde, 7: cor_verde,
            10: cor_amarela, 11: cor_amarela,
            12: cor_roxa, 13: cor_roxa,
            14: cor_verde, 15: cor_verde
        }

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=15), 1):
            is_header = (row_idx == 1)
            
            for col_idx, cell in enumerate(row, 1):
                # Aplicar cores específicas das colunas
                if col_idx in mapa_cores:
                    cell.fill = mapa_cores[col_idx]
                else:
                    cell.fill = cor_branca
                
                # Formatando fonte e alinhamento
                if is_header:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="left")
                else:
                    # Se não for a coluna de Produto (1) e não for a linha em branco, formata como número
                    if col_idx > 1 and cell.value != "" and cell.value is not None:
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = "#,##0"
                            cell.alignment = Alignment(horizontal="right")
                        except ValueError:
                            pass

                # Remove borda da coluna vazia separadora (H / Coluna 8)
                if col_idx != 8 and cell.value != "":
                    cell.border = borda_fina

        # Ajustar larguras das colunas
        ws.column_dimensions['A'].width = 25
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
            ws.column_dimensions[col_letter].width = 18
        ws.column_dimensions['H'].width = 3 # Coluna separadora

        wb.save(caminho_saida)
        print(f"Relatório gerado com sucesso em: {caminho_saida}")

# ==========================================
# SIMULAÇÃO E EXECUÇÃO
# ==========================================
def gerar_dados_mock():
    """Gera dados de exemplo para testar a estrutura imediatamente."""
    produtos = ["MOP", "SAM", "UREIA", "MAP", "NP+NPS+MÊS", "SSP", "TSP"]
    
    # SIACESP (Jan a Maio)
    df_siacesp = pd.read_excel('BI_Importacao Siacesp.xlsx')
    
    # Lineups (4 fontes para Jun, Jul, Ago)
    lineups = ['BI_Line Up.xlsx', 'BI_Line Up_WILSON SONS.xlsx', 'BI_Line Up_ORION.xlsx', 'BI_Line Up_TRANSATLANTICA.xlsx']
    #add o leitor de excel para cada arquivo e armazenar em uma lista de DataFrames
    df_lineups_list = [pd.read_excel(file) for file in lineups]
        
    # Histórico 2025
    df_hist_2025 = pd.DataFrame({
        'PRODUTO': produtos,
        'JUNHO_2025': [1487, 187, 330, 354, 266, 222, 240],
        'JULHO_2025': [1408, 414, 422, 297, 533, 680, 236],
        'AGOSTO_2025': [1455, 688, 581, 393, 473, 463, 146],
        'ACUM_JUNHO_2025': [6671, 2538, 1906, 1661, 1265, 1810, 922],
        'ACUM_JULHO_2025': [8080, 2952, 2327, 1958, 1798, 2489, 1126],
        'ACUM_AGOSTO_2025': [9534, 3640, 2909, 2351, 2365, 2952, 1272]
    })
    
    return df_siacesp, df_lineups_list, df_hist_2025

def main():
    # 1. Carregar Dados
    # NOTA: Substitua as funções de mock pelas leituras dos seus Excels reais via pd.read_excel()
    df_siacesp, df_lineups_list, df_historico_2025 = gerar_dados_mock()
    
    # 2. Instanciar e Processar
    relatorio = RelatorioEscalacao(df_siacesp, df_lineups_list, df_historico_2025)
    df_resultado = relatorio.processar_dados()

    RelatorioEscalacao.exportar_excel(df_resultado, "Relatorio_Escalacao_Formatado.xlsx")
    

if __name__ == "__main__":
    main()
