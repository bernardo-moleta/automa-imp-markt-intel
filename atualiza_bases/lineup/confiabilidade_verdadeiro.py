import json
import pandas as pd
import warnings
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# Ignora avisos de tipos mistos que podem aparecer ao ler arquivos CSV grandes
warnings.filterwarnings("ignore")


class ComparadorEscalacao:
    def __init__(self, siacesp_path, data_posicao_alvo, mes_etb_alvo=5):
        """
        Inicializa a classe centralizando a base do SIACESP.
        """
        self.df_siacesp = pd.read_excel(siacesp_path)
        self.data_posicao_alvo = pd.to_datetime(data_posicao_alvo).date()
        self.mes_etb_alvo = mes_etb_alvo

        # 1. Converte a coluna de data do SIACESP (ajuste 'Z_PERÍODO' se o nome da coluna for diferente)
        self.df_siacesp["Z_PERÍODO"] = pd.to_datetime(
            self.df_siacesp["Z_PERÍODO"], errors="coerce"
        )

        # 2. Filtra o SIACESP para o mês e ano desejados (Ex: Maio de 2026)
        df_siacesp_filtrado = self.df_siacesp[
            (self.df_siacesp["Z_PERÍODO"].dt.month == self.mes_etb_alvo)
            & (self.df_siacesp["Z_PERÍODO"].dt.year == 2026)
        ]

        # 3. Converte o volume para numérico e agrega APENAS a base filtrada
        df_siacesp_filtrado["VOLUME"] = pd.to_numeric(
            df_siacesp_filtrado["VOLUME"], errors="coerce"
        ).fillna(0)
        self.siacesp_agg = df_siacesp_filtrado.groupby(
            "Y_PRODUTO PARA", as_index=False
        )["VOLUME"].sum()
        self.siacesp_agg.rename(columns={"VOLUME": "Volume_SIACESP"}, inplace=True)

    def processar_lineup(self, nome, caminho, col_data_pos, col_etb, col_volume):
        """
        Lê, limpa, filtra e compara um arquivo individual de Escalação (Line Up).
        """
        df = pd.read_excel(caminho)

        # Tratamento das colunas de data
        df[col_data_pos] = pd.to_datetime(df[col_data_pos], errors="coerce")
        # format='mixed' e dayfirst=False ajudam com padrões irregulares como MM/DD/YYYY
        df[col_etb] = pd.to_datetime(
            df[col_etb], errors="coerce", format="mixed", dayfirst=False
        )

        # Aplica o filtro de Data da Posição e Mês do ETB
        df_filtrado = df[df[col_data_pos].dt.date == self.data_posicao_alvo]
        df_mes = df_filtrado[df_filtrado[col_etb].dt.month == self.mes_etb_alvo]

        # Converte a coluna de volume customizada e agrega os dados
        df_mes[col_volume] = pd.to_numeric(df_mes[col_volume], errors="coerce").fillna(
            0
        )
        lineup_agg = df_mes.groupby("Y_PRODUTO PARA", as_index=False)[col_volume].sum()
        lineup_agg.rename(columns={col_volume: f"Volume_LineUp"}, inplace=True)

        # Realiza um outer merge para garantir que produtos exclusivos de uma base também apareçam
        comparacao = pd.merge(
            lineup_agg, self.siacesp_agg, on="Y_PRODUTO PARA", how="outer"
        ).fillna(0)

        # Calcula a discrepância
        comparacao["Discrepancia"] = (
            comparacao[f"Volume_LineUp"] - comparacao["Volume_SIACESP"]
        )

        # Organiza as maiores discrepâncias no topo e adiciona a identificação da origem
        comparacao = comparacao.sort_values(
            by="Discrepancia", key=abs, ascending=False
        ).reset_index(drop=True)
        comparacao.insert(0, "Fonte Line Up", nome)

        # Arredonda os valores decimais para manter o relatório limpo
        comparacao["Volume_LineUp"] = comparacao["Volume_LineUp"].round(2)
        comparacao["Volume_SIACESP"] = comparacao["Volume_SIACESP"].round(2)
        comparacao["Discrepancia"] = comparacao["Discrepancia"].round(2)

        return comparacao


def aplicar_estilo_planilha(ws, titulo):
    """
    Formata as células do Excel com padrão visual (cores corporativas e zebrado).
    """
    header_fill = PatternFill(
        start_color="2c3e50", end_color="2c3e50", fill_type="solid"
    )
    header_font = Font(name="Arial", size=11, bold=True, color="ffffff")
    zebra_fill = PatternFill(
        start_color="f2f4f5", end_color="f2f4f5", fill_type="solid"
    )
    white_fill = PatternFill(
        start_color="ffffff", end_color="ffffff", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin", color="d3d3d3"),
        right=Side(style="thin", color="d3d3d3"),
        top=Side(style="thin", color="d3d3d3"),
        bottom=Side(style="thin", color="d3d3d3"),
    )

    # Insere e pinta o título superior
    ws.insert_rows(1, 2)
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="2c3e50")

    # Itera sobre todas as colunas para aplicar estilos e auto-ajustar larguras
    for col_idx, col in enumerate(
        ws.iter_cols(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column)
    ):
        max_length = 0
        for row_idx, cell in enumerate(col):
            if row_idx == 0:  # Linha de Cabeçalho
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:  # Dados
                cell.fill = zebra_fill if row_idx % 2 == 0 else white_fill
                cell.font = Font(name="Arial", size=10)

                # Alinha números à direita e textos à esquerda
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0.00"
                else:
                    cell.alignment = Alignment(horizontal="left")

            cell.border = thin_border

            # Checa o tamanho para ajuste da coluna
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = max_length + 2
        ws.column_dimensions[col[0].column_letter].width = (
            adjusted_width if adjusted_width > 12 else 12
        )

    ws.freeze_panes = (
        "A4"  # Congela as linhas de cabeçalho e título ao rolar para baixo
    )


def exportar_dados_para_html(
    excel_path="Analise_Comparativa_Escalacao.xlsx",
    df_siacesp_bruto=None,
    dict_lineups=None,
    config_sop=None,
):
    """
    dict_lineups: Dicionário com nomes das fontes e seus respectivos arquivos (ou DataFrames).
        Ex: {
            "ORION": "BI_Line Up_ORION.xlsx",
            "TRANSATLANTICA": "BI_Line Up_TRANSATLANTICA.xlsx",
            "WSONS": "BI_Line Up_WILSON SONS.xlsx",
            "UNIMAR": "BI_Line Up.xlsx"
        }
    config_sop: Dicionário definindo quais fontes usar em cada mês de projeção S&OP (1, 2 e 3).
        Ex: {
            1: ["ORION", "TRANSATLANTICA"], # Junho usa a média destas duas
            2: ["WSONS"],                   # Julho usa apenas esta
            3: ["ORION", "UNIMAR"]          # Agosto usa a média destas duas
        }
    """
    print("\nExportando dados dinâmicos para o Dashboard HTML...")
    try:
        xls = pd.ExcelFile(excel_path)

        def ler_aba(nome_aba):
            if nome_aba in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=nome_aba, skiprows=2)
                return df.fillna(0).to_dict(orient="records")
            return []

        if isinstance(df_siacesp_bruto, str):
            df_siacesp_bruto = pd.read_excel(df_siacesp_bruto)

        if df_siacesp_bruto is not None:
            for col in df_siacesp_bruto.select_dtypes(
                include=["datetime64", "datetimetz"]
            ).columns:
                df_siacesp_bruto[col] = df_siacesp_bruto[col].astype(str)

        # --- CONSOLIDADOR DE MÚLTIPLOS LINE UPS ---
        lineup_raw_consolidado = []
        if dict_lineups:
            for nome_fonte, caminho_ou_df in dict_lineups.items():
                if isinstance(caminho_ou_df, str):
                    df_temp = pd.read_excel(caminho_ou_df)
                else:
                    df_temp = caminho_ou_df.copy()

                for col in df_temp.select_dtypes(
                    include=["datetime64", "datetimetz"]
                ).columns:
                    df_temp[col] = df_temp[col].astype(str)

                df_temp["FONTE_DADOS"] = nome_fonte  # Carimba de onde veio a informação
                lineup_raw_consolidado.extend(
                    df_temp.fillna(0).to_dict(orient="records")
                )

        # Configuração padrão de segurança (se não for enviada, usa todas as fontes para todos os meses)
        if not config_sop and dict_lineups:
            todas_fontes = list(dict_lineups.keys())
            config_sop = {1: todas_fontes, 2: todas_fontes, 3: todas_fontes}

        datasets = {
            "resumo": ler_aba("Resumo Consolidado"),
            "unimar": ler_aba("Comp_UNIMAR"),
            "transatlantica": ler_aba("Comp_TRANSATLANTICA"),
            "wilson": ler_aba("Comp_WILSON_SONS"),
            "orion": ler_aba("Comp_ORION"),
            "siacesp_base": ler_aba("Base_SIACESP_Agregada"),
            "siacesp_raw": (
                df_siacesp_bruto.fillna(0).to_dict(orient="records")
                if df_siacesp_bruto is not None
                else []
            ),
            "lineup_raw": lineup_raw_consolidado,
            "config_sop": config_sop or {},
        }

        with open("dados_dashboard.js", "w", encoding="utf-8") as f:
            f.write("const dashboardData = " + json.dumps(datasets, default=str) + ";")

        print("✅ Arquivo 'dados_dashboard.js' atualizado com Suporte Multi-Lineup!")
    except Exception as e:
        print(f"Erro ao exportar dados para HTML: {e}")


def main():
    # 1. Instanciar a classe base configurando Data da Posição (2026-06-11) e Mês ETB (Maio)
    # NOTA: Ajuste o nome dos arquivos nos parâmetros conforme estão salvos localmente
    comparador = ComparadorEscalacao("BI_Importacao Siacesp.xlsx", "2026-06-11", 5)

    # 2. Processar cada base com suas colunas de volume específicas
    comp_unimar = comparador.processar_lineup(
        nome="UNIMAR",
        caminho="BI_Line Up.xlsx",
        col_data_pos="DATA_POSIÇÃO",
        col_etb="ETB",
        col_volume="TONNAGE",
    )

    comp_trans = comparador.processar_lineup(
        nome="TRANSATLANTICA",
        caminho="BI_Line Up_TRANSATLANTICA.xlsx",
        col_data_pos="DATA POSIÇÃO",
        col_etb="ETB",
        col_volume="QUANTITY",
    )

    comp_wilson = comparador.processar_lineup(
        nome="WILSON SONS",
        caminho="BI_Line Up_WILSON SONS.xlsx",
        col_data_pos="DATA POSIÇÃO",
        col_etb="ETB",
        col_volume="WEIGHT",
    )

    comp_orion = comparador.processar_lineup(
        nome="ORION",
        caminho="BI_Line Up_ORION.xlsx",
        col_data_pos="DATA POSIÇÃO",
        col_etb="Z_ETB",
        col_volume="Qtty",
    )

    # 3. Concatenar as bases e extrair as Top 20 maiores divergências num contexto consolidado
    resumo = pd.concat([comp_unimar, comp_trans, comp_wilson, comp_orion])
    resumo = (
        resumo[resumo["Discrepancia"] != 0]
        .sort_values(by="Discrepancia", key=abs, ascending=False)
        .head(20)
    )

    # 4. Construção do Arquivo Excel usando OpenPyXL
    wb = Workbook()

    # Aba 1: Resumo Global
    ws_summary = wb.active
    ws_summary.title = "Resumo Consolidado"
    for r in dataframe_to_rows(resumo, index=False, header=True):
        ws_summary.append(r)
    aplicar_estilo_planilha(
        ws_summary, "Top 20 Discrepâncias - Múltiplas Fontes (Maio 2026)"
    )

    # Aba 2: UNIMAR
    ws_unimar = wb.create_sheet(title="Comp_UNIMAR")
    for r in dataframe_to_rows(comp_unimar, index=False, header=True):
        ws_unimar.append(r)
    aplicar_estilo_planilha(ws_unimar, "Comparação: Line Up UNIMAR vs SIACESP")

    # Aba 3: Transatlântica
    ws_trans = wb.create_sheet(title="Comp_TRANSATLANTICA")
    for r in dataframe_to_rows(comp_trans, index=False, header=True):
        ws_trans.append(r)
    aplicar_estilo_planilha(ws_trans, "Comparação: Line Up TRANSATLANTICA vs SIACESP")

    # Aba 4: Wilson Sons
    ws_wilson = wb.create_sheet(title="Comp_WILSON_SONS")
    for r in dataframe_to_rows(comp_wilson, index=False, header=True):
        ws_wilson.append(r)
    aplicar_estilo_planilha(ws_wilson, "Comparação: Line Up WILSON SONS vs SIACESP")

    ws_orion = wb.create_sheet(title="Comp_ORION")
    for r in dataframe_to_rows(comp_orion, index=False, header=True):
        ws_orion.append(r)
    aplicar_estilo_planilha(ws_orion, "Comparação: Line Up ORION vs SIACESP")

    # Aba 5: Base de Segurança SIACESP (apenas os volumes agrupados originais)
    ws_siacesp = wb.create_sheet(title="Base_SIACESP_Agregada")
    for r in dataframe_to_rows(
        comparador.siacesp_agg.sort_values(by="Volume_SIACESP", ascending=False),
        index=False,
        header=True,
    ):
        ws_siacesp.append(r)
    aplicar_estilo_planilha(ws_siacesp, "SIACESP - Volume Total por Produto")

    # 5. Salva o resultado final
    wb.save("Analise_Comparativa_Escalacao.xlsx")
    print("Relatório 'Analise_Comparativa_Escalacao.xlsx' gerado com sucesso!")

    # ==========================================

    # ADICIONE A CHAMADA DA FUNÇÃO AQUI:
    # Certifique-se de que o nome do arquivo bate com o que você acabou de salvar
    exportar_dados_para_html(
        excel_path="Analise_Comparativa_Escalacao.xlsx",
        df_siacesp_bruto="BI_Importacao Siacesp.xlsx",
        dict_lineups={
            "ORION": "BI_Line Up_ORION.xlsx",
            "TRANSATLANTICA": "BI_Line Up_TRANSATLANTICA.xlsx",
            "WSONS": "BI_Line Up_WILSON SONS.xlsx",
            "UNIMAR": "BI_Line Up.xlsx",
        },
        config_sop={
            1: ["UNIMAR", "TRANSATLANTICA", "ORION"],
            2: ["UNIMAR", "TRANSATLANTICA", "ORION"],
            3: ["ORION"],
        },
    )


# Executa o pipeline
if __name__ == "__main__":
    main()
